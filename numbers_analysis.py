import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
import numpy as np
# Define the lists
            # "Big": [22, 18, 29, 7, 28, 12, 35, 3, 26, 0, 32, 15, 19, 4, 21, 2, 25],
            # "Orph": [1, 20, 14, 31, 9, 17, 34, 6],
            # "Small": [27, 13, 36, 11, 30, 8, 23, 10, 5, 24, 16, 33],

big = [22, 18, 29, 7, 28, 12, 35, 3, 26, 0, 32, 15, 19, 4, 21, 2, 25]
orphan = [1, 20, 14, 31, 9, 17, 34, 6]
small = [27, 13, 36, 11, 30, 8, 23, 10, 5, 24, 16, 33]
WHEEL_LAYOUT = [0, 32, 15, 19, 4, 21, 2, 25, 17, 34, 6, 27, 13, 36, 11, 30, 8, 23, 10, 5, 24, 16, 33, 1, 20, 14, 31, 9, 22, 18, 29, 7, 28, 12, 35, 3, 26]
NUMBER_TO_POSITION = {num: idx for idx, num in enumerate(WHEEL_LAYOUT)}
def wheel_distance(num1, num2):

    # <class 'numpy.float64'> <class 'numpy.float64'>
    if num1 == 1500 or num2 == 1500:
        return 1500
    pos1 = NUMBER_TO_POSITION[num1]
    pos2 = NUMBER_TO_POSITION[num2]
    dist = min((pos1 - pos2) % len(WHEEL_LAYOUT), (pos2 - pos1) % len(WHEEL_LAYOUT))
    max_dist = len(WHEEL_LAYOUT) // 2
    return min(dist, max_dist)
# read all csv files from csv_files folder
merged_data = pd.DataFrame()
for file in os.listdir('csv_files'):

    filename = os.path.join('csv_files', file)  
    if filename.endswith('.csv'):
        # Read the CSV file into a DataFrame
        df = pd.read_csv(filename)
        # do it with pandas
        df['Next Number'] = df['Number'].shift(-1)
        #df.dropna(inplace=True)
        df.fillna(value=1500, inplace=True)
        df['Distance'] = df.apply(lambda x: wheel_distance(x['Number'], x['Next Number']), axis=1)

        merged_data = pd.concat([merged_data, df], axis=1)
        # Save DataFrame to Excel
        df.to_excel("data.xlsx", index=False)
        
        # Load the workbook and select the active sheet
        workbook = openpyxl.load_workbook("data.xlsx")
        sheet = workbook.active

        # light orange for small, light red for big, light blue for orphan
        # Define the colors
        big_fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")  # Red for big
        small_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Orange for small
        orphan_fill = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")
        # there is only one column in the excel file
        # Iterate over the rows and apply colors based on the lists
        
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value in big:
                    cell.fill = big_fill
                elif cell.value in small:
                    cell.fill = small_fill
                elif cell.value in orphan:
                    cell.fill = orphan_fill

        # for each next number calculate the distance to the previous number, create new column

        # Save the modified workbook
        filename = os.path.splitext(filename)[0]
    

        workbook.save(filename + ".xlsx")
merged_data.to_excel("merged_data.xlsx", index=False)
# merge all excel files into one 
# Load the first workbook
workbook = openpyxl.load_workbook("merged_data.xlsx")
sheet = workbook.active

# light orange for small, light red for big, light blue for orphan
# Define the colors
big_fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")  # Red for big
small_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Orange for small
orphan_fill = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")
# there is only one column in the excel file
# Iterate over the rows and apply colors based on the lists
# Number	Next Number	Distance	Number	Next Number	Distance	Number	Next Number	Distance	Number	Next Number	Distance	Number	Next Number	Distance	Number	Next Number	Distance	Number	Next Number	Distance	Number	Next Number	Distance	Number	Next Number	Distance	Number	Next Number	Distance	Number	Next Number	Distance	Number	Next Number	Distance	Number	Next Number	Distance	Number	Next Number	Distance

columns_to_color = [1, 4, 7, 10, 13, 16, 19, 22, 25, 28, 31, 34, 37,40]

for col in columns_to_color:
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col, max_col=col):
        for cell in row:
            if cell.value in big:
                cell.fill = big_fill
            elif cell.value in small:
                cell.fill = small_fill
            elif cell.value in orphan:
                cell.fill = orphan_fill

# Save the modified workbook
workbook.save("merged_data.xlsx")